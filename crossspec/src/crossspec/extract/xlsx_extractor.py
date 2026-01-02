"""XLSX extractor."""

from __future__ import annotations

from pathlib import Path
from typing import Iterable, List, Optional

from openpyxl import load_workbook

from crossspec.claims import Authority
from crossspec.config import XlsxConfig, XlsxTableConfig
from crossspec.extract.base import ExtractedClaim, Extractor


class XlsxExtractor(Extractor):
    def __init__(self, path: Path, authority: Authority, config: XlsxConfig) -> None:
        self.path = path
        self.authority = authority
        self.config = config

    def extract(self) -> Iterable[ExtractedClaim]:
        workbook = load_workbook(self.path, data_only=True)
        for table in self.config.tables:
            sheet = workbook[table.sheet]
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                continue
            headers = [str(value).strip() if value is not None else "" for value in rows[0]]
            header_index = {header: idx for idx, header in enumerate(headers) if header}
            for row_index, row in enumerate(rows[1:], start=2):
                text_lines: List[str] = []
                columns_snapshot = {}
                for header, idx in header_index.items():
                    value = row[idx] if idx < len(row) else None
                    if value is None:
                        continue
                    text_value = str(value)
                    columns_snapshot[header] = text_value
                    if header in table.text_columns:
                        text_lines.append(f"{header}: {text_value}")
                if not text_lines:
                    continue
                authority = self._authority_for_row(table, columns_snapshot) or self.authority
                provenance = {
                    "sheet": table.sheet,
                    "row": row_index,
                    "columns_snapshot": columns_snapshot,
                }
                yield ExtractedClaim(
                    text_raw="\n".join(text_lines),
                    source_type="xlsx",
                    source_path=str(self.path),
                    authority=authority,
                    provenance=provenance,
                )

    @staticmethod
    def _authority_for_row(
        table: XlsxTableConfig,
        columns_snapshot: dict,
    ) -> Optional[Authority]:
        authority_map = table.authority_by or {}
        for column_name, mapping in authority_map.items():
            value = columns_snapshot.get(column_name)
            if value is None:
                continue
            mapped = mapping.get(str(value))
            if mapped:
                return Authority(mapped)
        return None
